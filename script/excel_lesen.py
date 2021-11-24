import openpyxl as op

import openpyxl as op

#Lädt das Excel File und schreibt die Werte des gesamten Files eine Liste
def readSrcFile(ExcelSrc, srcList):
    wb = op.load_workbook(ExcelSrc,data_only=True) # lädt das File
    ws = wb.worksheets[0]
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    for value in ws.iter_rows(min_row=1, values_only=True):
            if value[0] is not None:
                srcList.append(value)

#Lädt das Excel File und schreibt die Werte einer bestimmten Spalte in eine Liste
def readSrcFileColumn(srcFile, dstList, srcColumn): # srcFile = Das ExcelFile | dstList = die Zielvariable | srcColumn = Index der srcSpalte
    wb = op.load_workbook(srcFile,data_only=True) # lädt das File
    ws = wb.worksheets[0]
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    for value in ws.iter_rows(min_row=1, values_only=True):
            if value[0] is not None:
                dstList.append(value[srcColumn])


